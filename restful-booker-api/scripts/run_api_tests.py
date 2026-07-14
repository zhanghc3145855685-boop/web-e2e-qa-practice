"""
Execute Restful-Booker API cases and write results to Excel.

Reads case definitions aligned with generate_excel.py / Postman collection,
calls the live API, then writes:
  ../data/api-execution-results.xlsx

Sheets:
  - ExecutionResults: per-case actual status, pass/fail, response snippet
  - Summary: totals and pass rate

Usage:
  py run_api_tests.py
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUT = DATA_DIR / "api-execution-results.xlsx"

BASE = "https://restful-booker.herokuapp.com"
TIMEOUT = 30

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")


def style_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = max(12, min(56, length + 2))


def call(method: str, path: str, **kwargs) -> requests.Response:
    return requests.request(method, BASE + path, timeout=TIMEOUT, **kwargs)


def run() -> Path:
    results: list[dict] = []
    token = ""
    booking_id: int | None = None
    json_headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    # API-001 Ping
    r = call("GET", "/ping")
    ok = r.status_code == 201 and "Created" in r.text
    results.append(_row("API-001", "GET", "/ping", 201, r, ok, "body contains Created"))

    # API-002 Auth
    r = call("POST", "/auth", headers=json_headers, json={"username": "admin", "password": "password123"})
    try:
        token = r.json().get("token", "")
    except Exception:
        token = ""
    ok = r.status_code == 200 and bool(token)
    results.append(_row("API-002", "POST", "/auth", 200, r, ok, f"token={token[:8]}..." if token else "missing token"))

    # API-003 Get all
    r = call("GET", "/booking", headers={"Accept": "application/json"})
    ok = False
    note = ""
    try:
        body = r.json()
        ok = r.status_code == 200 and isinstance(body, list) and len(body) > 0 and "bookingid" in body[0]
        if ok and booking_id is None:
            booking_id = int(body[0]["bookingid"])
        note = f"count={len(body) if isinstance(body, list) else 'n/a'}"
    except Exception as e:
        note = str(e)
    results.append(_row("API-003", "GET", "/booking", 200, r, ok, note))

    # API-004 Get by id (existing)
    read_id = booking_id or 1
    r = call("GET", f"/booking/{read_id}", headers={"Accept": "application/json"})
    ok = False
    note = ""
    try:
        b = r.json()
        ok = (
            r.status_code == 200
            and "firstname" in b
            and "lastname" in b
            and "bookingdates" in b
            and str(b["bookingdates"].get("checkin", "")).count("-") >= 2
        )
        note = f"id={read_id}; firstname={b.get('firstname')}"
    except Exception as e:
        note = str(e)
    results.append(_row("API-004", "GET", f"/booking/{read_id}", 200, r, ok, note))

    # API-005 Create
    create_payload = {
        "firstname": "ApiTest",
        "lastname": "DayTwo",
        "totalprice": 150,
        "depositpaid": True,
        "bookingdates": {"checkin": "2026-07-01", "checkout": "2026-07-05"},
        "additionalneeds": "Breakfast",
    }
    r = call("POST", "/booking", headers=json_headers, json=create_payload)
    ok = False
    note = ""
    try:
        body = r.json()
        ok = (
            r.status_code == 200
            and "bookingid" in body
            and body.get("booking", {}).get("firstname") == "ApiTest"
            and body.get("booking", {}).get("totalprice") == 150
        )
        if ok:
            booking_id = int(body["bookingid"])
        note = f"bookingId={body.get('bookingid')}"
    except Exception as e:
        note = str(e)
    results.append(_row("API-005", "POST", "/booking", 200, r, ok, note))

    # API-006 PUT
    put_payload = {
        "firstname": "Updated",
        "lastname": "FullName",
        "totalprice": 300,
        "depositpaid": False,
        "bookingdates": {"checkin": "2026-08-01", "checkout": "2026-08-10"},
        "additionalneeds": "Dinner",
    }
    r = call(
        "PUT",
        f"/booking/{booking_id}",
        headers={**json_headers, "Cookie": f"token={token}"},
        json=put_payload,
    )
    ok = False
    note = ""
    try:
        b = r.json()
        ok = r.status_code == 200 and b.get("firstname") == "Updated" and b.get("totalprice") == 300
        note = f"firstname={b.get('firstname')}; totalprice={b.get('totalprice')}"
    except Exception as e:
        note = str(e)
    results.append(_row("API-006", "PUT", f"/booking/{booking_id}", 200, r, ok, note))

    # API-007 PATCH
    r = call(
        "PATCH",
        f"/booking/{booking_id}",
        headers={**json_headers, "Cookie": f"token={token}"},
        json={"firstname": "Patched", "additionalneeds": "LateCheckout"},
    )
    ok = False
    note = ""
    try:
        b = r.json()
        ok = (
            r.status_code == 200
            and b.get("firstname") == "Patched"
            and b.get("lastname") == "FullName"
            and b.get("additionalneeds") == "LateCheckout"
        )
        note = f"firstname={b.get('firstname')}; lastname={b.get('lastname')}"
    except Exception as e:
        note = str(e)
    results.append(_row("API-007", "PATCH", f"/booking/{booking_id}", 200, r, ok, note))

    # API-008 DELETE
    r = call(
        "DELETE",
        f"/booking/{booking_id}",
        headers={"Content-Type": "application/json", "Cookie": f"token={token}"},
    )
    ok = r.status_code == 201 and "Created" in r.text
    results.append(_row("API-008", "DELETE", f"/booking/{booking_id}", 201, r, ok, r.text.strip()))

    # API-009 404
    r = call("GET", "/booking/99999999", headers={"Accept": "application/json"})
    ok = r.status_code == 404
    results.append(_row("API-009", "GET", "/booking/99999999", 404, r, ok, r.text.strip()[:80]))

    # API-010 invalid token
    r = call(
        "DELETE",
        "/booking/1",
        headers={"Content-Type": "application/json", "Cookie": "token=invalid_token_day2"},
    )
    ok = r.status_code == 403
    results.append(_row("API-010", "DELETE", "/booking/1", 403, r, ok, r.text.strip()[:80]))

    # API-011 bad auth
    r = call(
        "POST",
        "/auth",
        headers=json_headers,
        json={"username": "admin", "password": "wrong_password"},
    )
    ok = False
    note = ""
    try:
        body = r.json()
        ok = r.status_code == 200 and body.get("reason") == "Bad credentials"
        note = json.dumps(body, ensure_ascii=False)
    except Exception as e:
        note = str(e)
    results.append(_row("API-011", "POST", "/auth", 200, r, ok, note))

    return write_excel(results)


def _row(
    case_id: str,
    method: str,
    endpoint: str,
    expected: int,
    resp: requests.Response,
    passed: bool,
    note: str,
) -> dict:
    snippet = (resp.text or "")[:240].replace("\n", " ")
    return {
        "CaseID": case_id,
        "Method": method,
        "Endpoint": endpoint,
        "ExpectedStatus": expected,
        "ActualStatus": resp.status_code,
        "Result": "通过" if passed else "失败",
        "ElapsedMs": int(resp.elapsed.total_seconds() * 1000),
        "ChecksNote": note,
        "ResponseSnippet": snippet,
        "ExecutedAt": datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %z"),
    }


def write_excel(results: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "ExecutionResults"
    headers = [
        "CaseID",
        "Method",
        "Endpoint",
        "ExpectedStatus",
        "ActualStatus",
        "Result",
        "ElapsedMs",
        "ChecksNote",
        "ResponseSnippet",
        "ExecutedAt",
    ]
    style_header(ws, headers)
    for item in results:
        ws.append([item[h] for h in headers])
        result_cell = ws.cell(ws.max_row, 6)
        result_cell.fill = PASS_FILL if item["Result"] == "通过" else FAIL_FILL
    autosize(ws)

    total = len(results)
    passed = sum(1 for r in results if r["Result"] == "通过")
    failed = total - passed
    rate = f"{(passed / total * 100):.1f}%" if total else "N/A"

    summary = wb.create_sheet("Summary")
    style_header(summary, ["Metric", "Value"])
    for row in [
        ("BaseUrl", BASE),
        ("Total", total),
        ("Passed", passed),
        ("Failed", failed),
        ("PassRate", rate),
        ("ExecutedAt", datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %z")),
        ("Tool", "Python requests + openpyxl (mirrors Postman collection)"),
    ]:
        summary.append(list(row))
    autosize(summary)

    wb.save(OUT)
    print(f"Total={total} Passed={passed} Failed={failed} PassRate={rate}")
    print(f"Wrote {OUT}")
    return OUT


if __name__ == "__main__":
    run()
