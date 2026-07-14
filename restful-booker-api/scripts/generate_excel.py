"""
Generate Restful-Booker API test Excel workbooks (industry-style data-driven layout).

Outputs under ../data/:
  - api-test-cases.xlsx   : EnvConfig / TestCases / TestData sheets
  - (execution results are written by run_api_tests.py)

Usage:
  py generate_excel.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUT = DATA_DIR / "api-test-cases.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def autosize(ws, min_width: int = 12, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


ENV_ROWS = [
    ("baseUrl", "https://restful-booker.herokuapp.com", "API 基础地址"),
    ("username", "admin", "Auth 用户名（演示站公开账号）"),
    ("password", "password123", "Auth 密码（演示站公开账号）"),
    ("accept", "application/json", "默认 Accept"),
    ("contentType", "application/json", "默认 Content-Type"),
]

# Columns aligned with common API case sheets used in QA teams
CASE_HEADERS = [
    "CaseID",
    "Module",
    "Title",
    "Priority",
    "Type",
    "Method",
    "Endpoint",
    "DataRef",
    "AuthRequired",
    "Headers",
    "ExpectedStatus",
    "ExpectedChecks",
    "PostmanRequest",
    "Preconditions",
    "Steps",
]

CASES = [
    (
        "API-001",
        "Health",
        "Ping 健康检查",
        "P0",
        "正向",
        "GET",
        "/ping",
        "",
        "N",
        "",
        201,
        "body contains Created",
        "01 GET Ping Health Check",
        "服务可访问",
        "1. GET {{baseUrl}}/ping 2. 检查状态码与响应体",
    ),
    (
        "API-002",
        "Auth",
        "正确凭据获取 token",
        "P0",
        "正向",
        "POST",
        "/auth",
        "TD-AUTH-01",
        "N",
        "Content-Type: application/json",
        200,
        "json.token 非空字符串",
        "02 POST Auth Create Token",
        "无",
        "1. POST /auth 使用 TD-AUTH-01 2. 保存 token 供后续用例",
    ),
    (
        "API-003",
        "Booking",
        "获取全部 booking id",
        "P0",
        "正向",
        "GET",
        "/booking",
        "",
        "N",
        "Accept: application/json",
        200,
        "array 非空且元素含 bookingid",
        "03 GET All Bookings",
        "库中至少存在一条 booking",
        "1. GET /booking 2. 校验数组结构",
    ),
    (
        "API-004",
        "Booking",
        "按 id 查询单个 booking",
        "P0",
        "正向",
        "GET",
        "/booking/{id}",
        "TD-BOOK-READ",
        "N",
        "Accept: application/json",
        200,
        "含 firstname/lastname/totalprice/bookingdates；日期 YYYY-MM-DD",
        "04 GET Booking By Id",
        "已知有效 bookingId（可由 API-003/005 产生）",
        "1. GET /booking/{id} 2. 校验关键字段",
    ),
    (
        "API-005",
        "Booking",
        "新建 booking",
        "P0",
        "正向",
        "POST",
        "/booking",
        "TD-BOOK-CREATE",
        "N",
        "Content-Type/Accept: application/json",
        200,
        "返回 bookingid；booking.firstname/lastname/totalprice 与请求一致",
        "05 POST Create Booking",
        "无",
        "1. POST /booking 使用 TD-BOOK-CREATE 2. 记录 bookingId",
    ),
    (
        "API-006",
        "Booking",
        "全量更新 booking (PUT)",
        "P0",
        "正向",
        "PUT",
        "/booking/{id}",
        "TD-BOOK-PUT",
        "Y",
        "Cookie: token={{token}}",
        200,
        "返回体字段与 PUT 请求体一致",
        "06 PUT Update Booking",
        "已有 token（API-002）与 bookingId（API-005）",
        "1. PUT /booking/{id} 2. 校验全量字段",
    ),
    (
        "API-007",
        "Booking",
        "部分更新 booking (PATCH)",
        "P1",
        "正向",
        "PATCH",
        "/booking/{id}",
        "TD-BOOK-PATCH",
        "Y",
        "Cookie: token={{token}}",
        200,
        "firstname/additionalneeds 更新；其余字段保留",
        "07 PATCH Partial Update Booking",
        "已有 token 与 bookingId",
        "1. PATCH /booking/{id} 2. 校验部分字段",
    ),
    (
        "API-008",
        "Booking",
        "删除 booking",
        "P0",
        "正向",
        "DELETE",
        "/booking/{id}",
        "",
        "Y",
        "Cookie: token={{token}}",
        201,
        "body contains Created（该 API 删除成功响应约定）",
        "08 DELETE Booking",
        "已有 token 与可删除的 bookingId",
        "1. DELETE /booking/{id} 2. 校验 201",
    ),
    (
        "API-009",
        "Booking",
        "查询不存在的 booking id",
        "P1",
        "负向",
        "GET",
        "/booking/99999999",
        "",
        "N",
        "Accept: application/json",
        404,
        "Not Found",
        "09 NEG GET Nonexistent Booking Id",
        "无",
        "1. GET /booking/99999999 2. 期望 404",
    ),
    (
        "API-010",
        "Booking",
        "错误 token 删除 booking",
        "P0",
        "负向",
        "DELETE",
        "/booking/1",
        "TD-NEG-TOKEN",
        "Y",
        "Cookie: token=invalid_token_day2",
        403,
        "Forbidden",
        "10 NEG DELETE With Invalid Token",
        "无",
        "1. DELETE /booking/1 带错误 token 2. 期望 403",
    ),
    (
        "API-011",
        "Auth",
        "错误密码获取 token",
        "P1",
        "负向",
        "POST",
        "/auth",
        "TD-AUTH-BAD",
        "N",
        "Content-Type: application/json",
        200,
        "json.reason == Bad credentials",
        "11 NEG POST Auth Bad Credentials",
        "无",
        "1. POST /auth 使用错误密码 2. 校验 reason",
    ),
]

DATA_HEADERS = [
    "DataRef",
    "Description",
    "firstname",
    "lastname",
    "totalprice",
    "depositpaid",
    "checkin",
    "checkout",
    "additionalneeds",
    "username",
    "password",
    "token",
    "Notes",
]

TEST_DATA = [
    (
        "TD-AUTH-01",
        "合法登录凭据",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "admin",
        "password123",
        "",
        "来自官方文档公开账号",
    ),
    (
        "TD-AUTH-BAD",
        "错误密码",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "admin",
        "wrong_password",
        "",
        "负向：期望 Bad credentials",
    ),
    (
        "TD-BOOK-CREATE",
        "新建订单数据",
        "ApiTest",
        "DayTwo",
        150,
        True,
        "2026-07-01",
        "2026-07-05",
        "Breakfast",
        "",
        "",
        "",
        "与 Postman 05 请求体一致",
    ),
    (
        "TD-BOOK-PUT",
        "全量更新数据",
        "Updated",
        "FullName",
        300,
        False,
        "2026-08-01",
        "2026-08-10",
        "Dinner",
        "",
        "",
        "",
        "与 Postman 06 请求体一致",
    ),
    (
        "TD-BOOK-PATCH",
        "部分更新数据",
        "Patched",
        "",
        "",
        "",
        "",
        "",
        "LateCheckout",
        "",
        "",
        "",
        "仅更新 firstname + additionalneeds",
    ),
    (
        "TD-BOOK-READ",
        "读取用占位",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "bookingId 运行时由 Create/List 注入",
    ),
    (
        "TD-NEG-TOKEN",
        "非法 token",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "invalid_token_day2",
        "负向删除鉴权",
    ),
]


def build() -> Path:
    wb = Workbook()

    ws_env = wb.active
    ws_env.title = "EnvConfig"
    style_header(ws_env, ["Key", "Value", "Description"])
    for row in ENV_ROWS:
        ws_env.append(list(row))
    autosize(ws_env)

    ws_cases = wb.create_sheet("TestCases")
    style_header(ws_cases, CASE_HEADERS)
    for row in CASES:
        ws_cases.append(list(row))
    for row in ws_cases.iter_rows(min_row=2, max_row=ws_cases.max_row):
        for cell in row:
            cell.alignment = WRAP
    autosize(ws_cases, max_width=56)
    ws_cases.column_dimensions["O"].width = 40

    ws_data = wb.create_sheet("TestData")
    style_header(ws_data, DATA_HEADERS)
    for row in TEST_DATA:
        ws_data.append(list(row))
    autosize(ws_data)

    ws_map = wb.create_sheet("FieldDictionary")
    style_header(ws_map, ["Sheet", "Column", "Meaning"])
    dict_rows = [
        ("TestCases", "CaseID", "用例唯一编号"),
        ("TestCases", "DataRef", "关联 TestData 行，实现数据与步骤分离"),
        ("TestCases", "ExpectedStatus", "期望 HTTP 状态码"),
        ("TestCases", "ExpectedChecks", "断言要点（状态码外的字段/业务校验）"),
        ("TestCases", "PostmanRequest", "对应 Postman Collection 中的请求名称"),
        ("TestData", "DataRef", "被 TestCases.DataRef 引用"),
        ("EnvConfig", "baseUrl", "切换环境时只改此表"),
    ]
    for row in dict_rows:
        ws_map.append(list(row))
    autosize(ws_map)

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
